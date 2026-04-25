"""
بوت تيليجرام - البحث عن الإيميل بالرقم القومي
مزاولة المهنة - دور مايو 2026
"""

import logging
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, CommandHandler, filters, ContextTypes

# =============================================
# ضع هنا التوكن بتاع البوت من BotFather
BOT_TOKEN = "8760936418:AAHZ5qzI7uOPRybhwHnsMPVu1psLag45C3E"

# ضع هنا المسار الكامل لملف الإكسيل
EXCEL_FILE = "مزاولةالمهنه_دور_مايو_2026.xlsx"
# =============================================

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)


def load_data(filepath: str) -> dict:
    """تحميل بيانات الإكسيل في dictionary"""
    data = {}
    wb = load_workbook(filepath, read_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            national_id = row[2]   # العمود C - الرقم القومي
            email = row[6]         # العمود G - الإيميل
            name = row[1]          # العمود B - الاسم
            
            if national_id and email:
                # تحويل الرقم القومي لنص للتأكد من التطابق
                key = str(national_id).strip()
                data[key] = {
                    "email": str(email).strip(),
                    "name": str(name).strip() if name else ""
                }
    
    wb.close()
    logger.info(f"تم تحميل {len(data)} سجل من الملف")
    return data


# تحميل البيانات عند بدء البوت
print("⏳ جاري تحميل بيانات الإكسيل...")
DATA = load_data(EXCEL_FILE)
print(f"✅ تم تحميل {len(DATA)} سجل بنجاح")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """رسالة الترحيب"""
    await update.message.reply_text(
        "👋 أهلاً بك!\n\n"
        "🔍 أرسل لي رقمك القومي (14 رقم)\n"
        "وهرد عليك بالإيميل المسجل في امتحان مزاولة المهنة دور مايو 2026"
    )


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """معالجة الرسائل والبحث بالرقم القومي"""
    text = update.message.text.strip()
    
    # التحقق إن الرسالة أرقام فقط
    if not text.isdigit():
        await update.message.reply_text(
            "⚠️ من فضلك أرسل الرقم القومي فقط (أرقام فقط، 14 رقم)"
        )
        return
    
    # التحقق من طول الرقم القومي
    if len(text) != 14:
        await update.message.reply_text(
            f"⚠️ الرقم القومي لازم يكون 14 رقم\n"
            f"اللي بعتته {len(text)} رقم"
        )
        return
    
    # البحث في البيانات
    if text in DATA:
        record = DATA[text]
        email = record["email"]
        name = record["name"]
        
        # إخفاء جزء من الإيميل لحماية الخصوصية
        masked = mask_email(email)
        
        await update.message.reply_text(
            f"✅ تم العثور على بياناتك!\n\n"
            f"👤 الاسم: {name}\n"
            f"📧 الإيميل المسجل:\n`{masked}`\n\n"
            f"_إذا كان الإيميل غلط، تواصل مع الجهة المختصة_",
            parse_mode="Markdown"
        )
        logger.info(f"بحث ناجح - رقم قومي: {text[:4]}**********")
    else:
        await update.message.reply_text(
            "❌ الرقم القومي ده مش موجود في قاعدة البيانات\n\n"
            "تأكد من:\n"
            "• صحة الرقم القومي\n"
            "• إنك مسجل في امتحان دور مايو 2026"
        )
        logger.info(f"بحث فاشل - رقم قومي: {text[:4]}**********")


def mask_email(email: str) -> str:
    """إخفاء جزء من الإيميل للخصوصية"""
    if "@" not in email:
        return email
    
    local, domain = email.split("@", 1)
    
    if len(local) <= 3:
        masked_local = local[0] + "*" * (len(local) - 1)
    else:
        visible = max(2, len(local) // 3)
        masked_local = local[:visible] + "*" * (len(local) - visible)
    
    return f"{masked_local}@{domain}"


def main():
    """تشغيل البوت"""
    print(f"🤖 جاري تشغيل البوت...")
    
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    
    # إضافة الأوامر والمعالجات
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    print("✅ البوت شغال! اضغط Ctrl+C لإيقافه")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
